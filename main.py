import pandas as pd  # type: ignore
import openpyxl  # type: ignore


# Essa função lê somente a coluna B e extrai todos os dados
def lerFormularioLimpo():
    colunab = 'B'
    dados_limpos = pd.read_excel('limpar_dados.xlsx', usecols=colunab)
    return dados_limpos

def inserirDadosEmCelulas(dados_limpos, planilhaTransportes='Formulario new.xlsx', planilha_final='Formulario - xxx.xlsx'):
    # Carregar a planilha de destino    
    wb = openpyxl.load_workbook(planilhaTransportes)
    ws = wb.active  # Use a aba ativa

    # Lista de células específicas onde os dados serão inseridos
    celulas = ['AA10', 'W10', 'S13', 'W13', 'S17', 'S10', 'U17', 'Y20', 'S20', 'AC20']

    # Inserir os dados nas células especificadas
    for i, celula in enumerate(celulas):
        if i < len(dados_limpos):  # Verifica se existem dados suficientes
            ws[celula] = dados_limpos.iloc[i, 0]  # Insere o valor na célula pulando uma pra baixo.

    # Salvar a nova planilha 
    wb.save(planilha_final)
    print(f"Dados inseridos e salvos em '{planilha_final}'.")

    return wb  # Retornar o workbook para uso em outras funções

# Função para gerar a mensagem personalizada
def gerarMensagem(wb):
    ws = wb.active  # Acessar a aba ativa do workbook
    endereco = ws['W13'].value  # Pegar o valor da célula W13
    nome = ws['S20'].value  # Supondo que o nome está em W10 (ajuste se necessário)

    # Validar se os valores foram encontrados
    if not endereco or not nome:
        print("Erro: Não foi possível encontrar todos os dados necessários para gerar a mensagem.")
        return

    # Mensagem personalizada
    mensagem = f"""
    Solicitação:
    
    Nota de trabalho: Envio via correios
    
    OS: xxxx - Endereço de envio: {endereco}
    A/C: {nome}
    
    Prazo para a entrega é de 5 dias úteis
    """
    print(mensagem)

# Chamando as funções
dados = lerFormularioLimpo()
if dados is not None:
    workbook = inserirDadosEmCelulas(dados)
    gerarMensagem(workbook)
